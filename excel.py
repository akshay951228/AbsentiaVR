import string
from openpyxl import load_workbook
import formulas
import argparse
import pandas as pd
import _pickle as pkl
import openpyxl


class ExcelProcessing():
    def __init__(self,template_path):
        wb = load_workbook(template_path)
        sheets = wb.sheetnames
        self.ws = wb[sheets[0]]
        self.input_dict ={}   #inputs columns {cell_id:column_name}
        self.output_dict ={}  #output columns {cell_id:column_name}
        self.mappers={}     # {column_name:cell_id}
        self.r_mappers ={}  #{cell_id:column_name}
        self.excel_formulas ={}    # {'column_name': '=IF(LEN(G3)>0,TEXT(G3,"00,000"),"")'}
        self.process_formulas = {}  # present this is useless
        self.formula_params ={}  # {'column_name': ['column_name'],}
        self.xl_formt_params = {} # {'column_name': ['cell_id'],}
        self.column_names = []  #[A,B,.....AA,..]
        for x in range(self.ws.max_column):
            if x < 26:
                self.column_names.append(string.ascii_uppercase[x % 26])
            else:
                self.column_names.append(self.column_names[int(x/26 - 1)] + string.ascii_uppercase[int(x % 26)])
   
    def formula_parser(self,formulea):
        return formulas.Parser().ast(formulea)[1].compile()

    def get_info_from_template(self):
        # processing all info updating inputs,outputs,mappers 
        for _col in self.column_names:
            col_name = self.ws[_col][1].value
            if self.ws[_col][0].value.lower().startswith("input"):
                self.input_dict[_col] = col_name
            elif self.ws[_col][0].value.lower().startswith("output"):
                self.output_dict[_col] = col_name
            is_formula = self.ws[_col][2].value
            if str(is_formula).startswith("="):
                self.excel_formulas[col_name] = str(is_formula)
            self.mappers[col_name] = _col
            self.r_mappers[_col] = col_name

        for key, val in self.excel_formulas.items():
            func = self.formula_parser(val)
            params=list(func.inputs)
            updated_params = []
            for _each_p in params:
                col = ''.join([s for s in _each_p if not s.isdigit()])
                col_name = self.r_mappers[col]
                self.xl_formt_params[key] = _each_p
            updated_params.append(col_name)
            self.process_formulas[key] = func 
            self.formula_params[key] = updated_params
    def process_input_columns(self,_raw_data,rd_column):
        temp_input_dict = self.input_dict.copy()
        count = 0
        df = None
        while(bool(temp_input_dict)):
            col = self.column_names[count]
            col_name = temp_input_dict[col]
            if col_name in rd_column:
                if df is None:
                    df =pd.DataFrame({col_name:_raw_data[col_name]})
                    updated_col = col
                else:
                   df[str(col_name)] = _raw_data[col_name]
                   updated_col = col
                del temp_input_dict[col]
            count = count+1
        return df,count







def main_processing(template_path,csv_path,output_path,output_pkl_path,sample=0):
    #input_paths
    i_df = pd.read_csv(csv_path)
    template_path = template_path
    intermediate_output_path = output_path
    if sample!=0:
        i_df = i_df.head(1000)
    res_col = i_df.to_dict(orient='li')
    _input_column_name_csv=list(i_df)
    len_raw_data  = i_df.shape[0]

    writer = pd.ExcelWriter(intermediate_output_path, engine='xlsxwriter')

    ep_instance = ExcelProcessing(template_path)
    ep_instance.get_info_from_template()
    df,count = ep_instance.process_input_columns(res_col,_input_column_name_csv)


    df.to_excel(writer, sheet_name='Sheet1', index=False)
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # output columns updating
    for cnt in range(count,len(ep_instance.column_names)-1):
        _col = ep_instance.column_names[cnt+1]
        _name = ep_instance.r_mappers[_col]
        try:
            o_formale_params = ep_instance.xl_formt_params[_name]
            xl_formula = ep_instance.excel_formulas[_name]
            # print(_col,_name,o_formale_params,xl_formula)
        except Exception as e:
            continue
        u_cell_name= ''.join([s for s in o_formale_params if not s.isdigit()])
        worksheet.write(_col+str(1), _name)
        for i in range(len_raw_data):
            u_cell_name= ''.join([s for s in o_formale_params if not s.isdigit()])
            u_cell_name = u_cell_name+str(i+2)
            # print(_col,_name,o_formale_params,xl_formula,u_cell_name)
            updated_formula = xl_formula.replace(o_formale_params,u_cell_name)
            worksheet.write_formula(_col+str(i+2), updated_formula)
    writer.save()

    #saving the pickel file 
    with open(output_pkl_path, 'wb') as handle:
        pkl.dump(ep_instance.output_dict, handle)


def postprocess(interm_excel,pkl_path,output_path):
    required_col = []
    with open(pkl_path, 'rb') as handle:
        output_dict = pkl.load(handle)
    for key,val in output_dict.items():
        required_col.append(val)
    
    excel_data_df = pd.read_excel(interm_excel, sheet_name='Sheet1')
    xls_col_name = list(excel_data_df.columns)
    
    update_columns = []
    for _n in required_col:
        if _n in xls_col_name:
            update_columns.append(_n)
    final_df = excel_data_df[update_columns]   
    final_df.to_csv (output_path, index = False, header=True)

    import ipdb; ipdb.set_trace()
if __name__ == "__main__":
        # ARGUMENTS --------------------------------------------------------------------------------------------------------
    parser = argparse.ArgumentParser(description='Excel processing')
    subparsers = parser.add_subparsers(title="subcommands", dest="subcommand")

    # ARGUMENTS: Processing the csv to intermediate excel file  -------------------------------------------------------------------------------
    process = subparsers.add_parser("process", help="Pre-process the dataset for its use.")
    process.add_argument("--template_path", type=str, required=True,
                                help="Path to the template file")
    process.add_argument("--csv_path", type=str, required=True,
                                help="Path to the unprocessed csv file")
    process.add_argument("--output_path", type=str, required=True,
                                help="Path to store file path")
    process.add_argument("--output_pkl_path", type=str, required=True,
                                help="Path to store pkl file")
    process.add_argument("--sample", type=int, required=False,default= 0,
                                help="Path to store pkl file")

    # ARGUMENTS: Post-processing  ----------------------------------------------------------------------------------------
    post_process = subparsers.add_parser("post_process", help="start the post processing")
    post_process.add_argument("--interm_xlsx", type=str, required=True,
                              help="Path to the intermediate file")
    post_process.add_argument("--pkl_path", type=str, required=True,
                              help="Path to the saved pkl path")
    post_process.add_argument("--output_path", type=str, required=True,
                                help="Path to store file path")

    args = parser.parse_args()

    # EXECUTE ----------------------------------------------------------------------------------------------------------
    try:
        if args.subcommand == "process":
            main_processing(
                template_path=args.template_path,
                csv_path=args.csv_path,
                output_path = args.output_path,
                output_pkl_path =  args.output_pkl_path,
                sample = args.sample,
            )
        elif args.subcommand == "post_process":
            postprocess(
                interm_excel=args.interm_xlsx,
                pkl_path = args.pkl_path,
                output_path=args.output_path,
           
            )
        else:
            print("invalid command")
    except Exception as e:
        print('Something went wrong',e )